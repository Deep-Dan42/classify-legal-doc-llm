"""
automacao-juridica-trf1 — Gerador de Relatório Excel

Gera relatório de oportunidades com 4 abas:
    1. Dados da Empresa (resumo)
    2. Serviços Identificados (1 linha por documento classificado)
    3. Oportunidades (gap analysis — PRODUTO PRINCIPAL)
    4. Triagem Pendente (itens REVISÃO, se houver)

Tema visual: escritório jurídico (laranja #D4700A, preto #1A1A1A, prata #C0C0C0)

Uso:
    from scripts.gerar_relatorio import gerar_relatorio_excel

    gerar_relatorio_excel(
        dados_empresa={...},
        servicos_identificados=[...],
        oportunidades=[...],
        triagem_pendente=[...],
        caminho_saida=Path("data/saida/relatorio.xlsx"),
    )
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.modelos import ItemRelatorio, Oportunidade


# ============================================
# TEMA VISUAL — ESCRITÓRIO JURÍDICO
# ============================================

_FONT_FAMILY = "Arial"

# Cores
_COR_LARANJA = "D4700A"
_COR_PRETO = "1A1A1A"
_COR_GRAFITE = "2D2D2D"
_COR_PRATA = "C0C0C0"
_COR_PRATA_CLARA = "E8E8E8"
_COR_BRANCO = "FFFFFF"

# Status
_COR_CONFIRMADO_BG = "E6F4E6"
_COR_CONFIRMADO_TX = "1A6B1A"
_COR_AUTOMATICO_BG = "F5E6D0"
_COR_AUTOMATICO_TX = "8B5E0A"
_COR_PENDENTE_BG = "FCE8E8"
_COR_PENDENTE_TX = "A33030"

# Estilos reutilizáveis
_HEADER_FILL = PatternFill("solid", fgColor=_COR_LARANJA)
_HEADER_FONT = Font(name=_FONT_FAMILY, bold=True, color=_COR_BRANCO, size=11)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

_CELL_FONT = Font(name=_FONT_FAMILY, size=10, color=_COR_PRETO)
_CELL_FONT_SMALL = Font(name=_FONT_FAMILY, size=9, color=_COR_GRAFITE)
_CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

_ZEBRA_FILL = PatternFill("solid", fgColor=_COR_PRATA_CLARA)

_THIN_BORDER = Border(
    left=Side(style="thin", color=_COR_PRATA),
    right=Side(style="thin", color=_COR_PRATA),
    top=Side(style="thin", color=_COR_PRATA),
    bottom=Side(style="thin", color=_COR_PRATA),
)

_TITLE_FONT = Font(name=_FONT_FAMILY, bold=True, size=14, color=_COR_PRETO)
_SUBTITLE_FONT = Font(name=_FONT_FAMILY, bold=True, size=11, color=_COR_GRAFITE)
_LABEL_FONT = Font(name=_FONT_FAMILY, size=10, color=_COR_PRATA)
_VALUE_FONT = Font(name=_FONT_FAMILY, bold=True, size=11, color=_COR_PRETO)


def _apply_header_row(ws, row: int, headers: list[str], widths: list[int]) -> None:
    """Aplica estilo de cabeçalho a uma linha."""
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _apply_data_row(ws, row: int, values: list, zebra: bool = False) -> None:
    """Aplica estilo de dados a uma linha."""
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = _CELL_FONT
        cell.alignment = _CELL_ALIGNMENT
        cell.border = _THIN_BORDER
        if zebra:
            cell.fill = _ZEBRA_FILL


def _apply_status_style(cell, status: str) -> None:
    """Aplica cor de status (Confirmado/Automático/Pendente)."""
    status_lower = status.lower()
    if "confirmado" in status_lower or "alta" in status_lower:
        cell.fill = PatternFill("solid", fgColor=_COR_CONFIRMADO_BG)
        cell.font = Font(name=_FONT_FAMILY, size=10, color=_COR_CONFIRMADO_TX, bold=True)
    elif "automático" in status_lower or "llm" in status_lower:
        cell.fill = PatternFill("solid", fgColor=_COR_AUTOMATICO_BG)
        cell.font = Font(name=_FONT_FAMILY, size=10, color=_COR_AUTOMATICO_TX, bold=True)
    elif "pendente" in status_lower or "revisão" in status_lower or "revisao" in status_lower:
        cell.fill = PatternFill("solid", fgColor=_COR_PENDENTE_BG)
        cell.font = Font(name=_FONT_FAMILY, size=10, color=_COR_PENDENTE_TX, bold=True)


# ============================================
# ABA 1: DADOS DA EMPRESA
# ============================================

def _criar_aba_empresa(wb: Workbook, dados: dict) -> None:
    """Cria aba com resumo da empresa."""
    ws = wb.active
    ws.title = "Dados da Empresa"
    ws.sheet_properties.tabColor = _COR_LARANJA

    # Título
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "Relatório de Oportunidades"
    title_cell.font = _TITLE_FONT

    ws.merge_cells("A2:D2")
    date_cell = ws["A2"]
    date_cell.value = f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    date_cell.font = _LABEL_FONT

    # Dados da empresa
    campos = [
        ("Empresa", dados.get("empresa", "")),
        ("CNPJ", dados.get("cnpj", "")),
        ("Cidade/UF", dados.get("cidade_uf_matriz", "")),
        ("Atividade Principal", dados.get("atividade_principal", "")),
        ("Atividade Secundária", dados.get("atividade_secundaria", "")),
        ("Capital Social", dados.get("capital_social", "")),
        ("Advogado", dados.get("advogado", "")),
    ]

    row = 4
    for label, value in campos:
        ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = _VALUE_FONT
        row += 1

    # Métricas resumo
    row += 1
    ws.cell(row=row, column=1, value="Resumo").font = _SUBTITLE_FONT
    row += 1
    metricas = [
        ("Processos analisados", dados.get("n_processos", 0)),
        ("Documentos classificados", dados.get("n_documentos", 0)),
        ("Serviços identificados", dados.get("n_identificados", 0)),
        ("Oportunidades", dados.get("n_oportunidades", 0)),
        ("Aguardando revisão", dados.get("n_revisao", 0)),
    ]
    for label, value in metricas:
        ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = _VALUE_FONT
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 45


# ============================================
# ABA 2: SERVIÇOS IDENTIFICADOS
# ============================================

def _criar_aba_servicos(wb: Workbook, itens: list[ItemRelatorio]) -> None:
    """Cria aba com serviços identificados por documento."""
    ws = wb.create_sheet("Serviços Identificados")
    ws.sheet_properties.tabColor = _COR_PRETO

    headers = [
        "Nº Processo", "Tipo Documento", "Tipo de Serviço",
        "Confiança", "Método", "Trecho Extraído", "Observação",
    ]
    widths = [28, 16, 35, 15, 12, 55, 40]

    _apply_header_row(ws, 1, headers, widths)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max(len(itens) + 1, 2)}"

    for i, item in enumerate(itens):
        row = i + 2
        zebra = i % 2 == 1

        values = [
            item.numero_processo,
            item.tipo_documento,
            item.tipo_de_servico,
            item.confianca_display(),
            item.metodo,
            item.trecho_extraido or "",
            item.observacao,
        ]
        _apply_data_row(ws, row, values, zebra=zebra)

        # Estilo de status na coluna Confiança
        _apply_status_style(ws.cell(row=row, column=4), item.confianca_display())

    # Altura das linhas com trecho
    for row in range(2, len(itens) + 2):
        ws.row_dimensions[row].height = 45


# ============================================
# ABA 3: OPORTUNIDADES
# ============================================

def _criar_aba_oportunidades(wb: Workbook, oportunidades: list[Oportunidade]) -> None:
    """Cria aba de oportunidades (PRODUTO PRINCIPAL)."""
    ws = wb.create_sheet("Oportunidades")
    ws.sheet_properties.tabColor = _COR_CONFIRMADO_TX  # verde

    headers = [
        "#", "Serviço Disponível", "Área Responsável",
        "Ramo de Atividade", "Regime Tributário",
        "Objeto da Tese", "Observação",
    ]
    widths = [5, 40, 20, 25, 25, 45, 30]

    _apply_header_row(ws, 1, headers, widths)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max(len(oportunidades) + 1, 2)}"

    for i, op in enumerate(oportunidades):
        row = i + 2
        zebra = i % 2 == 1

        values = [
            i + 1,
            op.servico_disponivel,
            op.area_responsavel,
            op.ramo_de_atividade,
            op.regime_tributario,
            op.objeto_da_tese,
            op.observacao,
        ]
        _apply_data_row(ws, row, values, zebra=zebra)


# ============================================
# ABA 4: TRIAGEM PENDENTE
# ============================================

def _criar_aba_triagem(wb: Workbook, pendentes: list[ItemRelatorio]) -> None:
    """Cria aba de triagem (itens REVISÃO)."""
    ws = wb.create_sheet("Triagem Pendente")
    ws.sheet_properties.tabColor = _COR_PENDENTE_TX  # vermelho

    headers = [
        "Nº Processo", "Tipo Documento", "Trecho Extraído",
        "Sugestão do Sistema", "Observação",
    ]
    widths = [28, 16, 60, 35, 40]

    _apply_header_row(ws, 1, headers, widths)
    ws.freeze_panes = "A2"

    if not pendentes:
        ws.cell(row=2, column=1, value="Nenhum item pendente de revisão.").font = _CELL_FONT
        ws.merge_cells("A2:E2")
        return

    for i, item in enumerate(pendentes):
        row = i + 2
        zebra = i % 2 == 1

        values = [
            item.numero_processo,
            item.tipo_documento,
            item.trecho_extraido or "",
            item.observacao or "",
            "",
        ]
        _apply_data_row(ws, row, values, zebra=zebra)

    for row in range(2, len(pendentes) + 2):
        ws.row_dimensions[row].height = 60


# ============================================
# GERADOR PRINCIPAL
# ============================================

def gerar_relatorio_excel(
    dados_empresa: dict,
    servicos_identificados: list[ItemRelatorio],
    oportunidades: list[Oportunidade],
    triagem_pendente: Optional[list[ItemRelatorio]] = None,
    caminho_saida: Optional[Path] = None,
) -> Path:
    """
    Gera relatório Excel com 4 abas.

    Args:
        dados_empresa: dict com metadados da empresa + métricas resumo.
        servicos_identificados: lista de ItemRelatorio (serviços classificados).
        oportunidades: lista de Oportunidade (gap analysis).
        triagem_pendente: lista de ItemRelatorio com confiança REVISÃO.
        caminho_saida: caminho do arquivo Excel. Se None, gera automaticamente.

    Returns:
        Path do arquivo gerado.
    """
    from core.config import settings

    triagem_pendente = triagem_pendente or []

    if caminho_saida is None:
        cnpj = dados_empresa.get("cnpj", "00000000000000")
        data = datetime.now().strftime("%Y%m%d")
        nome = f"relatorio_de_oportunidades_{cnpj}_{data}.xlsx"
        caminho_saida = settings.DATA_SAIDA_DIR / nome

    caminho_saida.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    _criar_aba_empresa(wb, dados_empresa)
    _criar_aba_servicos(wb, servicos_identificados)
    _criar_aba_oportunidades(wb, oportunidades)
    _criar_aba_triagem(wb, triagem_pendente)

    wb.save(str(caminho_saida))

    return caminho_saida